"""Query live PG schema and write Excel file with one sheet per table + Overview."""
from __future__ import annotations

from pathlib import Path
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "db_schema.xlsx"
DSN = "host=localhost port=5432 dbname=ran_dt user=ran password=ran"

RAN_TABLES = [
    "scene_snapshot",
    "gnb_config",
    "gnb_state",
    "ue_config",
    "ue_state",
    "position_history",
    "signal_history",
    "platform_events",
]

TABLE_PURPOSE = {
    "scene_snapshot":   "SceneIngestor 寫入的場景定義快照（來自 scene_config.json 或 ingest）",
    "gnb_config":       "gNB 靜態設定：頻率、功率、頻寬、啟停狀態",
    "gnb_state":        "gNB 即時位置快照（目前未大量使用，未來 gNB 可動時用）",
    "ue_config":        "UE 軌跡設定：waypoints、speed、loop（由 UEController.trajectory 寫入）",
    "ue_state":         "UE 即時快照：位置、serving cell、RSRP、SINR（由 SignalIngestor 更新）",
    "position_history": "UE 位置時序（S7 的 1Hz poller 會往這裡寫）",
    "signal_history":   "訊號時序：每筆 ingest 的 RSRP/SINR/rsrp_map 全保留",
    "platform_events":  "PlatformReporter 寫入的上行事件（目前只落 log，未來真的 POST 給平台）",
}


def fetch_schema():
    conn = psycopg2.connect(DSN)
    tables_data = []
    index_data = []
    with conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT t.table_name,
                       c.ordinal_position,
                       c.column_name,
                       c.data_type,
                       c.character_maximum_length,
                       c.is_nullable,
                       c.column_default,
                       CASE WHEN pk.column_name IS NOT NULL THEN 'PK' ELSE '' END AS is_pk,
                       CASE WHEN uq.column_name IS NOT NULL THEN 'UQ' ELSE '' END AS is_unique
                FROM information_schema.tables t
                JOIN information_schema.columns c ON c.table_name = t.table_name
                LEFT JOIN (
                    SELECT kcu.table_name, kcu.column_name
                    FROM information_schema.table_constraints tc
                    JOIN information_schema.key_column_usage kcu
                      ON kcu.constraint_name = tc.constraint_name
                     AND kcu.table_schema = tc.table_schema
                    WHERE tc.constraint_type = 'PRIMARY KEY'
                      AND tc.table_schema = 'public'
                ) pk ON pk.table_name = t.table_name AND pk.column_name = c.column_name
                LEFT JOIN (
                    SELECT kcu.table_name, kcu.column_name
                    FROM information_schema.table_constraints tc
                    JOIN information_schema.key_column_usage kcu
                      ON kcu.constraint_name = tc.constraint_name
                     AND kcu.table_schema = tc.table_schema
                    WHERE tc.constraint_type = 'UNIQUE'
                      AND tc.table_schema = 'public'
                ) uq ON uq.table_name = t.table_name AND uq.column_name = c.column_name
                WHERE t.table_schema = 'public'
                  AND c.table_schema = 'public'
                  AND t.table_name = ANY(%s)
                ORDER BY t.table_name, c.ordinal_position;
            """, (RAN_TABLES,))
            tables_data = cur.fetchall()

            cur.execute("""
                SELECT tablename, indexname, indexdef
                FROM pg_indexes
                WHERE schemaname = 'public'
                  AND tablename = ANY(%s)
                ORDER BY tablename, indexname;
            """, (RAN_TABLES,))
            index_data = cur.fetchall()

            # row counts
            counts = {}
            for t in RAN_TABLES:
                cur.execute(f"SELECT COUNT(*) FROM {t}")
                counts[t] = cur.fetchone()[0]
    conn.close()
    return tables_data, index_data, counts


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=13, bold=True, color="1F4E79")


def write_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")


def autosize(ws, min_w=10, max_w=60):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        mx = min_w
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is None:
                    continue
                length = max(len(str(line)) for line in str(cell.value).split("\n"))
                if length > mx:
                    mx = length
        ws.column_dimensions[letter].width = min(mx + 2, max_w)


def build_xlsx():
    rows, idx_rows, counts = fetch_schema()

    by_table: dict[str, list] = {t: [] for t in RAN_TABLES}
    for row in rows:
        tname = row[0]
        by_table.setdefault(tname, []).append(row)

    idx_by_table: dict[str, list] = {t: [] for t in RAN_TABLES}
    for r in idx_rows:
        idx_by_table.setdefault(r[0], []).append(r)

    wb = Workbook()

    # ---- Overview sheet ----
    ws = wb.active
    ws.title = "Overview"
    ws.cell(row=1, column=1, value="Omniver Platform — DB Schema").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Database: ran_dt  (postgres:16-alpine) · 8 tables · 由 main/apps/ran/models/*.py 定義")

    headers = ["Table", "Purpose", "Columns", "Rows", "Indexes"]
    write_header(ws, row=4, headers=headers)
    for i, t in enumerate(RAN_TABLES, start=5):
        ws.cell(row=i, column=1, value=t)
        ws.cell(row=i, column=2, value=TABLE_PURPOSE.get(t, ""))
        ws.cell(row=i, column=3, value=len(by_table[t]))
        ws.cell(row=i, column=4, value=counts.get(t, 0))
        ws.cell(row=i, column=5, value=len(idx_by_table[t]))
    for r in range(5, 5 + len(RAN_TABLES)):
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    autosize(ws)

    # ---- Per-table sheets ----
    col_headers = ["#", "Column", "Type", "Max Len", "Nullable", "Default", "Key"]
    for t in RAN_TABLES:
        ws = wb.create_sheet(title=t[:31])  # Excel limits sheet name to 31 chars
        ws.cell(row=1, column=1, value=f"{t}").font = TITLE_FONT
        ws.cell(row=2, column=1, value=TABLE_PURPOSE.get(t, ""))
        ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)

        write_header(ws, row=4, headers=col_headers)
        r = 5
        for _tname, ordinal, col, dtype, maxlen, nullable, default, is_pk, is_uq in by_table[t]:
            keys = []
            if is_pk:
                keys.append(is_pk)
            if is_uq:
                keys.append(is_uq)
            ws.cell(row=r, column=1, value=ordinal)
            ws.cell(row=r, column=2, value=col)
            ws.cell(row=r, column=3, value=dtype)
            ws.cell(row=r, column=4, value=maxlen if maxlen else "")
            ws.cell(row=r, column=5, value=nullable)
            ws.cell(row=r, column=6, value=default if default else "")
            ws.cell(row=r, column=7, value=",".join(keys))
            r += 1

        # Indexes section
        r += 1
        ws.cell(row=r, column=1, value="Indexes").font = Font(bold=True)
        r += 1
        write_header(ws, row=r, headers=["Name", "Definition"])
        r += 1
        for _t, iname, idef in idx_by_table[t]:
            ws.cell(row=r, column=1, value=iname)
            ws.cell(row=r, column=2, value=idef)
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
            r += 1

        autosize(ws, min_w=10, max_w=80)

    wb.save(OUT)
    print(f"[ok] wrote {OUT}  ({len(RAN_TABLES)} tables)")


if __name__ == "__main__":
    build_xlsx()
